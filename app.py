import io
import csv
from datetime import datetime, date, timedelta, time
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

TEMPLATE_XLSX = "Carepacks Tool_TEMPLATE_UPDATED.xlsx"

st.set_page_config(page_title="HP Care Pack Checker", layout="wide")

st.title("HP Care Pack Checker")
st.caption("Upload the exported warranty CSV, view the computed table, download an updated Excel template, and generate consolidated calendar reminders.")

# ---------------- Helpers ----------------

def _to_date(x):
    """Parse common HP export date formats + Excel serials."""
    if x is None:
        return pd.NaT

    if isinstance(x, (datetime, pd.Timestamp)):
        return pd.to_datetime(x).normalize()

    s = str(x).strip().strip('"')
    if not s:
        return pd.NaT

    # Excel serial date support (e.g., 44251)
    if s.isdigit():
        try:
            return pd.to_datetime(int(s), unit="D", origin="1899-12-30")
        except Exception:
            pass

    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return pd.to_datetime(datetime.strptime(s, fmt).date())
        except Exception:
            pass

    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def compute_table(import_df: pd.DataFrame, today: date) -> pd.DataFrame:
    """Compute the warranty end date table for UI display."""
    df = import_df.copy()
    df.columns = [c.strip() for c in df.columns]

    rename_map = {
        "Serial number": "Serial number",
        "Serial Number": "Serial number",
        "Product number": "Product number",
        "Product Number": "Product number",
        "Product name": "Product name",
        "Product Name": "Product name",
        "Coverage status": "Coverage status",
        "Coverage Status": "Coverage status",
        "Warranty start date": "Warranty start date",
        "Warranty Start Date": "Warranty start date",
        "Warranty end date": "Warranty end date",
        "Warranty End Date": "Warranty end date",
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    for col in [
        "Serial number",
        "Product number",
        "Product name",
        "Coverage status",
        "Warranty start date",
        "Warranty end date",
    ]:
        if col not in df.columns:
            df[col] = ""

    df["Warranty start date"] = df["Warranty start date"].apply(_to_date)
    df["Warranty end date"] = df["Warranty end date"].apply(_to_date)

    today_ts = pd.to_datetime(today)
    df["Days to warranty end"] = (df["Warranty end date"] - today_ts).dt.days

    def policy(days):
        if pd.isna(days):
            return ""
        if days > 90:
            return "Warranty active (>90d remaining)"
        if days >= 0:
            return "In window (≤90d before expiry)"
        if abs(days) <= 30:
            return "Expired (≤30d)"
        if abs(days) <= 730:
            return "Expired (>30d and ≤24m)"
        return "Expired (>24m)"

    df["Policy status"] = df["Days to warranty end"].apply(policy)

    def cp_type(pol):
        if pol in ("In window (≤90d before expiry)", "Expired (≤30d)"):
            return "Post-Warranty Full Support (DesignJet) – 1Y or 2Y"
        if pol == "Expired (>30d and ≤24m)":
            return "Return-to-Service/Support + Post-Warranty Full Support (must be purchased together)"
        if pol == "Warranty active (>90d remaining)":
            return "Not yet eligible for Post-Warranty (window opens at 90d before expiry)"
        if pol == "Expired (>24m)":
            return "Not eligible for RTS/Post-Warranty per policy"
        return ""

    df["Care Pack type (suggested)"] = df["Policy status"].apply(cp_type)

    def cp_pn(pol):
        if pol in ("In window (≤90d before expiry)", "Expired (≤30d)"):
            return "UB8U7PE (1Y) / UB8U8PE (2Y)"
        if pol == "Expired (>30d and ≤24m)":
            return 'U67TVE (Return to Support Service for DesignJet Midrange 36") + UB8U7PE (1Y PW)'
        return ""

    df["Care Pack part number(s)"] = df["Policy status"].apply(cp_pn)

    def cp_advice(pol):
        if pol in ("In window (≤90d before expiry)", "Expired (≤30d)"):
            return "Eligible now: choose 1Y or 2Y Post-Warranty; do not stack multiple PW packs."
        if pol == "Expired (>30d and ≤24m)":
            return "Eligible via reinstatement: buy Return-to-Service/Support + 1Y Post-Warranty together."
        if pol == "Warranty active (>90d remaining)":
            return "Too early for Post-Warranty: wait until within 90 days of warranty end."
        if pol == "Expired (>24m)":
            return "Not eligible under fixed Care Pack rules; consider refresh/contractual support."
        return ""

    df["Care Pack options (HP policy)"] = df["Policy status"].apply(cp_advice)

    out_cols = [
        "Serial number",
        "Product number",
        "Product name",
        "Coverage status",
        "Warranty start date",
        "Warranty end date",
        "Days to warranty end",
        "Policy status",
        "Care Pack type (suggested)",
        "Care Pack part number(s)",
        "Care Pack options (HP policy)",
    ]
    df = df[out_cols]

    df = df.sort_values(by=["Warranty end date", "Serial number"], ascending=[True, True], na_position="last")
    return df


def write_csv_lines_into_template(csv_bytes: bytes) -> bytes:
    """Paste raw CSV lines into the 'imported data csv' sheet column A, preserving formulas & formatting."""
    wb = load_workbook(TEMPLATE_XLSX, data_only=False)
    if "imported data csv" not in wb.sheetnames:
        raise ValueError("Template missing required sheet: 'imported data csv'")

    ws = wb["imported data csv"]

    text = csv_bytes.decode("utf-8-sig", errors="ignore")
    lines = [ln.rstrip("\r") for ln in text.split("\n") if ln.strip()]

    # Clear a safe range in col A without depending on max_row (formatting may extend it)
    max_clear = max(len(lines) + 50, 1000)
    for r in range(1, max_clear + 1):
        ws.cell(r, 1).value = None

    for i, line in enumerate(lines, start=1):
        ws.cell(i, 1).value = line

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _ics_escape(text: str) -> str:
    if text is None:
        return ""
    return (
        str(text)
        .replace("\\", "\\\\")
        .replace("\n", "\\n")
        .replace(",", "\\,")
        .replace(";", "\\;")
    )


def generate_ics_reminders_consolidated_by_day(
    result_df: pd.DataFrame,
    today: date,
    lead_days=(30, 15),
    event_time: time = time(9, 0),
    duration_minutes: int = 10,
    show_as_available: bool = True,
    tz_name: str = "Africa/Johannesburg",
    category_name: str = "Red Category",
    max_items_in_body: int = 200,
) -> bytes:
    """Create ONE calendar event per reminder date, listing all items due that day."""
    df = result_df.copy()
    if "Warranty end date" not in df.columns:
        return b""

    df["Warranty end date"] = pd.to_datetime(df["Warranty end date"], errors="coerce")
    today_ts = pd.to_datetime(today)

    # non-expired items only
    df = df[df["Warranty end date"].notna() & (df["Warranty end date"] >= today_ts)].copy()

    reminders = []
    for _, row in df.iterrows():
        w_end = row["Warranty end date"].date()
        for ld in lead_days:
            rd = w_end - timedelta(days=int(ld))
            if rd < today:
                continue
            reminders.append({
                "remind_date": rd,
                "lead_days": int(ld),
                "serial": str(row.get("Serial number", "")).strip(),
                "product": str(row.get("Product name", "")).strip(),
                "prodno": str(row.get("Product number", "")).strip(),
                "coverage": str(row.get("Coverage status", "")).strip(),
                "warranty_end": w_end,
            })

    if not reminders:
        return ("BEGIN:VCALENDAR\r\nVERSION:2.0\r\nPRODID:-//HP//Care Pack Checker//EN\r\nEND:VCALENDAR\r\n").encode("utf-8")

    rem_df = pd.DataFrame(reminders)

    tz = ZoneInfo(tz_name)
    now_utc = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    transp = "TRANSPARENT" if show_as_available else "OPAQUE"

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//HP//Care Pack Checker//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
    ]

    for remind_date, g in rem_df.groupby("remind_date"):
        g = g.sort_values(["lead_days", "warranty_end", "serial"]).reset_index(drop=True)

        local_start = datetime.combine(remind_date, event_time).replace(tzinfo=tz)
        local_end = local_start + timedelta(minutes=int(duration_minutes))

        dtstart_utc = local_start.astimezone(ZoneInfo("UTC")).strftime("%Y%m%dT%H%M%SZ")
        dtend_utc = local_end.astimezone(ZoneInfo("UTC")).strftime("%Y%m%dT%H%M%SZ")

        total = len(g)
        lead_set = ", ".join(map(str, sorted(g["lead_days"].unique())))
        summary = f"Warranty reminders ({lead_set}d): {total} item(s)"

        header = [
            f"Reminder date: {remind_date.isoformat()}",
            f"Items: {total}",
            f"Lead times included: {lead_set} day(s)",
            "",
            "List:",
        ]

        body = []
        shown = g.head(max_items_in_body)
        for _, r in shown.iterrows():
            body.append(
                f"- [{r['lead_days']}d] {r['product']} | SN: {r['serial']} | PN: {r['prodno']} | Ends: {r['warranty_end'].isoformat()} | {r['coverage']}"
            )
        if total > max_items_in_body:
            body.append(f"...and {total - max_items_in_body} more item(s).")

        description = "\n".join(header + body)

        uid = f"consolidated-{remind_date.strftime('%Y%m%d')}@hpcarepackchecker"

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_utc}",
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(description)}",
            "STATUS:CONFIRMED",
            f"TRANSP:{transp}",
            f"DTSTART:{dtstart_utc}",
            f"DTEND:{dtend_utc}",
            f"CATEGORIES:{_ics_escape(category_name)}",
            "COLOR:#FF0000",
            "X-APPLE-CALENDAR-COLOR:#FF0000",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


# ---------------- UI ----------------

uploaded = st.file_uploader(
    "Upload your warranty export (CSV)",
    type=["csv", "txt"],
    help="Upload the CSV export that contains serial number, product name, coverage status, warranty start/end dates, and product number.",
)

colA, colB = st.columns([2, 1])

with colB:
    today = st.date_input("Today (for day calculations)", value=date.today())

if uploaded is None:
    st.info("Upload a CSV file to view the calculated warranty end date table.")
    st.stop()

raw_bytes = uploaded.getvalue()
raw_text = raw_bytes.decode("utf-8-sig", errors="ignore")

# Detect delimiter (best-effort)
try:
    dialect = csv.Sniffer().sniff(raw_text[:5000])
except Exception:
    dialect = csv.excel

reader = csv.reader(io.StringIO(raw_text), dialect=dialect, skipinitialspace=True)
rows = list(reader)

if not rows:
    st.error("The uploaded file appears empty.")
    st.stop()

# If single-column raw lines, parse again
if len(rows[0]) == 1 and "," in rows[0][0]:
    reader2 = csv.reader(io.StringIO("\n".join([r[0] for r in rows])), skipinitialspace=True)
    rows = list(reader2)

header = [h.strip() for h in rows[0]]
data = rows[1:]

import_df = pd.DataFrame(data, columns=header)

for c in import_df.columns:
    import_df[c] = import_df[c].astype(str).str.strip().str.strip('"')

result_df = compute_table(import_df, today)

with colA:
    st.subheader("Warranty End Date Table")
    st.dataframe(
        result_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Warranty start date": st.column_config.DateColumn(format="DD-MMM-YYYY"),
            "Warranty end date": st.column_config.DateColumn(format="DD-MMM-YYYY"),
            "Days to warranty end": st.column_config.NumberColumn(format="%d"),
        },
    )

st.divider()

# Calendar reminders
st.subheader("Calendar Reminders")
st.write("Generate consolidated reminders: one 10-minute entry per day (for non-expired items), at 30 and 15 days before warranty end.")

create_cal = st.checkbox("Create consolidated calendar reminders (.ics)", value=False)

if create_cal:
    lead_times = st.multiselect(
        "Reminder lead times (days before warranty end)",
        options=[90, 60, 30, 15, 7],
        default=[30, 15],
    )

    reminder_time = st.time_input("Reminder time (SAST)", value=time(9, 0))

    show_free = st.checkbox("Show as Available (does not block calendar time)", value=True)

    category_choice = st.selectbox("Category (best-effort color in Outlook)", ["HP Care Packs", "Urgent", "Red", "Custom..."], index=0)

custom_category = ""
if category_choice == "Custom...":
    custom_category = st.text_input("Custom category name", value="HP Care Packs")

category = custom_category if category_choice == "Custom..." 
else category_choice

if lead_times:
        ics_bytes = generate_ics_reminders_consolidated_by_day(
            result_df,
            today,
            lead_days=tuple(sorted(lead_times, reverse=True)),
            event_time=reminder_time,
            duration_minutes=10,
            show_as_available=show_free,
            tz_name="Africa/Johannesburg",
            category_name=category,
        )

        st.download_button(
            "Download consolidated reminders (.ics)",
            data=ics_bytes,
            file_name="HP_CarePack_Reminders_CONSOLIDATED.ics",
            mime="text/calendar",
        )

        st.caption("Note: Event color depends on your calendar app. Outlook uses categories for colors; create a matching category for consistent red.")
else:
        st.warning("Select at least one lead time.")


# Excel download
st.subheader("Download Excel (Template Preserved)")
st.write("Download an updated Excel workbook. This keeps all formulas & formatting from the template and pastes your CSV into the 'imported data csv' tab as raw CSV lines.")

try:
    updated_bytes = write_csv_lines_into_template(raw_bytes)
    st.download_button(
        "Download updated Excel (template preserved)",
        data=updated_bytes,
        file_name="Carepacks Tool_UPDATED.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.error(f"Could not generate Excel output: {e}")

st.caption("Tip: Run locally with: streamlit run app.py")
